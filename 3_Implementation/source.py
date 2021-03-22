import pandas as pd
import openpyxl
# openpyxl load_workbook( ) function is used when you have to access an MS Excel file in openpyxl module
# this function is used to access excel file in current working directory
from openpyxl import load_workbook

# Reading data from excel sheet
read = pd.read_excel('Data.xlsx', sheet_name=['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5'])

# Creating empty dataframe
temp = []
df1 = pd.DataFrame()
df4 = pd.DataFrame()
df5 = pd.DataFrame()
df6 = pd.DataFrame()

# to read data given excelsheet in same directory
inp = pd.DataFrame(list(read.items()))

# reading different sheets from excel file
df4 = pd.DataFrame(inp[1][0].columns)
df6 = df6.append(df4)

# excluding common data in excel sheets
for i in range(1, 5):
    df5 = pd.DataFrame(inp[1][i].columns[3:10])
    df6 = df6.append(df5)

# updating data in Dataframe
df = pd.DataFrame(df6)
# for extention part
df3 = pd.DataFrame()
n = int(input('Enter no of inputs:-'))
count = 0

# this part is used to ask input data from terminal
for _ in range(n):
    temp1 = []
    data = int(input('Enter your ps no:-'))
    name = str(input('Enter the Name:-'))
    email = str(input('Enter the email:-'))
    temp1.append(data)
    temp1.append(name)
    temp1.append(email)
    temp.append(temp1)

# to extract required data into master sheet

for i in temp:
    data, name, email = i
    y = read['Sheet1']
    y = y[(y['PS number'] == data) & (y['Display Name'] == name) & (y['Official Email Address'] == email)]
# If input is invalid
    if len(y) == 0:
        print('Invalid input')
        print('Enter valid input')
# If given input is valid
    else:
        df = pd.DataFrame(y, columns=['SL#', 'PS number', 'Display Name', 'Official Email Address'])
        for i in read.keys():
            x = read[i]
            t = x[(x['PS number'] == data) & (x['Display Name'] == name) & (x['Official Email Address'] == email)]
            col = x.columns

# count no of colomuns from sheets to summary sheet
            for j in col:
                df[j] = t[j]
                count = count+1
                df3.at[i, 'No of columns'] = count
    df1 = df1.append(df)
df2 = df1.describe()
df3.at[1, 'Total column count'] = (len(df1.columns)*n)
book = load_workbook('Data.xlsx')
writer = pd.ExcelWriter('Data.xlsx', engine='openpyxl')
writer.book = book

# to write data into master sheet
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

# to create new sheets and write extracted data
df1.to_excel(writer, sheet_name='master', index=False)
df3.to_excel(writer, sheet_name='summary', index=False)
writer.save()
