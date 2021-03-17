import pandas as pd
from openpyxl import load_workbook

Sheet_1 = pd.read_excel(r'Data.xlsx', sheet_name=0)
Sheet_2 = pd.read_excel(r'Data.xlsx', sheet_name=1)
Sheet_3 = pd.read_excel(r'Data.xlsx', sheet_name=2)
Sheet_4 = pd.read_excel(r'Data.xlsx', sheet_name=3)
Sheet_5 = pd.read_excel(r'Data.xlsx', sheet_name=4)
# Mergining All sheets in one master sheet
Sheet_6 = Sheet_1.merge(Sheet_2, on="PS number", how="left")
Sheet_7 = Sheet_6.merge(Sheet_3, on="PS number", how="left")
Sheet_8 = Sheet_7.merge(Sheet_4, on="PS number", how="left")
Sheet_9 = Sheet_8.merge(Sheet_5, on="PS number", how="left")
# creating new excel file
Sheet_9.to_excel("Output.xlsx", index=True)
# a=input()
# merging the files

a6 = a1.merge(a2, on="PS number", how="left")
a7 = a6.merge(a3, on="PS number", how="left")
a8 = a7.merge(a4, on="PS number", how="left")
a9 = a8.merge(a5, on="PS number", how="left")

print(a9)

# creating new excel file
# a9.to_excel("Output.xlsx", index = False)
# variable_1 = int(input("Enter PS number : "))
# print(variable_1)
variable_1 = int(input("Enter PS number : "))

pf_variable = pd.DataFrame(a9, columns=['PS number', 'Display Name', 'Pin Code', 'Fone No.', 'Salary',
                                        'Official Email Address'])
#print(pf_variable)

pf_variable.set_index('PS number', inplace=True)

res = pf_variable.loc[variable_1]
print(res)

path = r"D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx"

book = load_workbook(path)
writer = pd.ExcelWriter(path, engine="openpyxl")
writer.book = book
res.to_excel(writer, sheet_name="mastersheet")

writer.save()
writer.close()
 load_workbook.book = book